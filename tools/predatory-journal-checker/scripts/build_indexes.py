#!/usr/bin/env python3
"""
build_indexes.py — generates the three local lookup indexes for the
Journal Evidence Checker (gurjas.org) from raw source spreadsheets.

Corrected version: fixes the two verified bugs found in the earlier
Node.js draft (Active/Inactive column was being ignored; discontinuation
reason codes assumed a category — "publication concerns" — that does not
exist in this edition of the Scopus list).

Run: python3 build_indexes.py
Reads from ./data/raw/, writes to ./out/
"""
import json
import hashlib
import datetime
import openpyxl

RAW = "data/raw"
OUT = "out"
SALT = "gurjas"
TODAY = datetime.date.today().isoformat()


def normalize_issn(raw):
    """Strip to [0-9X], validate mod-11 check digit, return 8-char ISSN or None."""
    if not raw:
        return None
    clean = "".join(ch for ch in str(raw).upper() if ch.isdigit() or ch == "X")
    if len(clean) != 8:
        return None
    weights = [8, 7, 6, 5, 4, 3, 2]
    total = sum(int(clean[i]) * weights[i] for i in range(7))
    remainder = total % 11
    check = 11 - remainder
    expected = "0" if check == 11 else ("X" if check == 10 else str(check))
    return clean if clean[7] == expected else None


def hash_issn(issn):
    return hashlib.sha256((SALT + issn).encode()).hexdigest()[:12]


# ---------------------------------------------------------------
# 1. SCOPUS — hashed index, Active/Inactive column honoured (Fix 1),
#    real discontinuation reason codes only (Fix 2)
# ---------------------------------------------------------------
print("Processing Scopus source list ...")
scopus_idx = {}
active_count = inactive_count = disc_count = 0

wb = openpyxl.load_workbook(f"{RAW}/ext_list_Jun_2026.xlsx", read_only=True, data_only=True)

ws = wb["Scopus Sources Jun. 2026"]
rows = ws.iter_rows(min_row=2, values_only=True)
header = ["Sourcerecord ID", "Source Title", "ISSN", "EISSN", "Active or Inactive"]
for row in rows:
    issn, eissn, status = row[2], row[3], row[4]
    code = "A" if str(status).strip() == "Active" else "I"
    if code == "A":
        active_count += 1
    else:
        inactive_count += 1
    p = normalize_issn(issn)
    e = normalize_issn(eissn)
    if p:
        scopus_idx[hash_issn(p)] = code
    if e:
        scopus_idx[hash_issn(e)] = code

ws = wb["Discontinued Titles Jun. 2026"]
# header row is row 2 (index 1); data starts row 3
rows = ws.iter_rows(min_row=3, values_only=True)
for row in rows:
    if row[0] is None and row[1] is None:
        continue
    issn, eissn, reason = row[2], row[3], row[5]
    reason_clean = str(reason or "").strip().lower()
    if reason_clean == "discontinuation":
        code = "D:DISC"
    elif reason_clean == "journal change policy":
        code = "D:JCP"
    else:
        code = "D:XX"
    p = normalize_issn(issn)
    e = normalize_issn(eissn)
    if p:
        scopus_idx[hash_issn(p)] = code
        disc_count += 1
    if e:
        scopus_idx[hash_issn(e)] = code
        disc_count += 1

sorted_scopus = dict(sorted(scopus_idx.items()))
scopus_out = {
    "meta": {
        "source": "Scopus Source List",
        "edition": "June 2026",
        "retrieved": TODAY,
        "unique_hash_entries": len(sorted_scopus),
        "active_issns_seen": active_count,
        "inactive_issns_seen": inactive_count,
        "discontinued_issns_seen": disc_count,
        "note": "Keys are salted SHA-256(first 12 hex chars) of normalised ISSNs. "
                "No journal titles are stored or shippable from this file, "
                "per Elsevier's redistribution terms.",
    },
    "data": sorted_scopus,
}
with open(f"{OUT}/scopus.idx.2026-07.json", "w") as f:
    json.dump(scopus_out, f, separators=(",", ":"))
print(f"  Scopus: {active_count} active + {inactive_count} inactive + "
      f"{disc_count} discontinued ISSN fields -> {len(sorted_scopus)} unique hashed entries")

# ---------------------------------------------------------------
# 2. ABDC — plain ISSN-keyed index (public list, safe to ship plainly)
# ---------------------------------------------------------------
print("Processing ABDC Journal Quality List ...")
abdc_idx = {}
wb2 = openpyxl.load_workbook(f"{RAW}/ABDC-JQL-2025-v1-260326.xlsx", read_only=True, data_only=True)
ws2 = wb2["2025 JQL"]
rows = ws2.iter_rows(min_row=9, values_only=True)  # header at row 8 (1-indexed)
abdc_count = 0
for row in rows:
    if row[1] is None:
        continue
    issn, eissn, rating = row[3], row[4], row[7]
    rating_clean = str(rating or "").strip()
    if rating_clean not in ("A*", "A", "B", "C"):
        continue
    p = normalize_issn(issn)
    e = normalize_issn(eissn)
    if p:
        abdc_idx[p] = rating_clean
        abdc_count += 1
    if e:
        abdc_idx[e] = rating_clean
        abdc_count += 1

sorted_abdc = dict(sorted(abdc_idx.items()))
abdc_out = {
    "meta": {
        "source": "ABDC Journal Quality List",
        "edition": "2025 (current to March 2026 review)",
        "retrieved": TODAY,
        "rows": len(sorted_abdc),
    },
    "data": sorted_abdc,
}
with open(f"{OUT}/abdc.json", "w") as f:
    json.dump(abdc_out, f, separators=(",", ":"))
print(f"  ABDC: {abdc_count} ISSN fields -> {len(sorted_abdc)} unique entries")

# ---------------------------------------------------------------
# 3. Manifest
# ---------------------------------------------------------------
manifest = {
    "version": "2026-07",
    "files": {
        "scopus": "scopus.idx.2026-07.json",
        "abdc": "abdc.json",
        "hijacked": "hijacked.json",
    },
    "excluded": {
        "Scopus_book_list_Q2.xlsx": "ISBN-keyed book data — out of scope for a journal-level "
                                     "predatory checker (v2)."
    },
    "counts": {
        "scopus_entries": len(sorted_scopus),
        "abdc_entries": len(sorted_abdc),
    },
}
with open(f"{OUT}/indexes-manifest.json", "w") as f:
    json.dump(manifest, f, indent=2)

print("Done. Files written to ./out/")
